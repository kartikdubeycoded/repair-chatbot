import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def export_to_excel(output_file='conversation_logs.xlsx'):
    """Export conversation logs to a formatted Excel file"""
    
    print(" Exporting conversation logs to Excel...\n")
    
    # Connect to database
    conn = sqlite3.connect('conversation_logs.db')
    
    # Query all conversations
    query = """
        SELECT 
            id as 'ID',
            timestamp as 'Date & Time',
            user_question as 'User Question',
            bot_response as 'Bot Response',
            sources as 'Sources Used',
            context_info as 'Context Type',
            response_time_seconds as 'Response Time (s)'
        FROM conversations
        ORDER BY timestamp DESC
    """
    
    df = pd.read_sql_query(query, conn)
    
    # Get statistics
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM conversations")
    total_conversations = cursor.fetchone()[0]
    
    cursor.execute("SELECT AVG(response_time_seconds) FROM conversations")
    avg_response_time = cursor.fetchone()[0] or 0
    
    cursor.execute("SELECT context_info, COUNT(*) FROM conversations WHERE context_info IS NOT NULL GROUP BY context_info")
    context_stats = cursor.fetchall()
    
    conn.close()
    
    # Create Excel file with pandas
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write main data
        df.to_excel(writer, sheet_name='Conversations', index=False)
        
        # Create statistics sheet
        stats_data = {
            'Metric': [
                'Total Conversations',
                'Average Response Time (seconds)',
                'Export Date'
            ],
            'Value': [
                total_conversations,
                f"{avg_response_time:.2f}",
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ]
        }
        stats_df = pd.DataFrame(stats_data)
        stats_df.to_excel(writer, sheet_name='Statistics', index=False)
        
        # Create context breakdown sheet if available
        if context_stats:
            context_data = {
                'Context Type': [ctx[0] for ctx in context_stats],
                'Count': [ctx[1] for ctx in context_stats]
            }
            context_df = pd.DataFrame(context_data)
            context_df.to_excel(writer, sheet_name='Context Breakdown', index=False)
    
    # Format the Excel file
    wb = load_workbook(output_file)
    
    # Format Conversations sheet
    ws_conv = wb['Conversations']
    format_conversations_sheet(ws_conv)
    
    # Format Statistics sheet
    ws_stats = wb['Statistics']
    format_statistics_sheet(ws_stats)
    
    # Format Context Breakdown sheet if it exists
    if 'Context Breakdown' in wb.sheetnames:
        ws_context = wb['Context Breakdown']
        format_context_sheet(ws_context)
    
    # Save formatted workbook
    wb.save(output_file)
    
    print(f" Export complete!")
    print(f" File saved: {output_file}")
    print(f" Total conversations: {total_conversations}")
    print(f"  Average response time: {avg_response_time:.2f}s\n")
    
    return output_file

def format_conversations_sheet(ws):
    """Format the Conversations sheet"""
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Set column widths
    column_widths = {
        'A': 8,   # ID
        'B': 20,  # Date & Time
        'C': 40,  # User Question
        'D': 60,  # Bot Response
        'E': 40,  # Sources
        'F': 20,  # Context Type
        'G': 15   # Response Time
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Format data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Color code by context type
            if cell.column == 6 and cell.value:  # Context Type column
                if 'filtered' in str(cell.value).lower():
                    cell.fill = PatternFill(start_color="E7F4E4", end_color="E7F4E4", fill_type="solid")
                elif 'clarification' in str(cell.value).lower():
                    cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    
    # Freeze header row
    ws.freeze_panes = 'A2'

def format_statistics_sheet(ws):
    """Format the Statistics sheet"""
    
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Format headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    
    # Format data
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(horizontal='center')

def format_context_sheet(ws):
    """Format the Context Breakdown sheet"""
    
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15

def auto_open_excel(filename):
    """Try to open the Excel file automatically"""
    import os
    import platform
    
    system = platform.system()
    
    try:
        if system == 'Windows':
            os.startfile(filename)
        elif system == 'Darwin':  # macOS
            os.system(f'open "{filename}"')
        else:  # Linux
            os.system(f'xdg-open "{filename}"')
        print(f" Opening {filename}...")
    except Exception as e:
        print(f"ℹ  Please open {filename} manually")

if __name__ == "__main__":
    # Export to Excel
    output_file = export_to_excel()
    
    # Try to open automatically
    auto_open_excel(output_file)